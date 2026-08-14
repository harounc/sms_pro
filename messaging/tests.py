from datetime import timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.messages import get_messages
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from accounts.models import Company, CompanyTransaction, User
from contacts.models import Contact, ContactGroup
from messaging.models import Campaign, Message, Sender
from messaging.services import send_sms_now
from messaging.sms_gateway import send_sms, validate_sms_configuration
from messaging.tasks import enqueue_due_scheduled_messages, send_message_task


def make_campaign_excel(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["phone", "name"])
    for row in rows:
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return SimpleUploadedFile(
        "campagne.xlsx",
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class SmsGatewayTests(TestCase):
    @override_settings(SMS_API_FROM="")
    def test_send_sms_requires_sender_configuration(self):
        result = send_sms("+2250700000000", "Test")

        self.assertFalse(result["success"])
        self.assertIn("SMS_API_FROM", result["error"])

    @override_settings(SMS_API_FROM="", SMS_API_KEY_BASE64="", SMS_API_CLIENT_ID="", SMS_API_CLIENT_SECRET="")
    def test_send_sms_accepts_explicit_sender_before_token_lookup(self):
        result = send_sms("+2250700000000", "Test", sender="TEST")

        self.assertFalse(result["success"])
        self.assertIn("SMS_API_KEY_BASE64", result["error"])

    @override_settings(
        SMS_API_FROM="",
        SMS_API_KEY_BASE64="",
        SMS_API_CLIENT_ID="client",
        SMS_API_CLIENT_SECRET="secret",
    )
    def test_sms_configuration_accepts_client_credentials_with_sender(self):
        self.assertIsNone(validate_sms_configuration(sender="TEST"))

    @override_settings(SMS_API_FROM="0100000000", SMS_API_KEY_BASE64="configured")
    @patch("messaging.sms_gateway._get_token")
    @patch("messaging.sms_gateway.requests.post")
    def test_send_sms_prefers_selected_sender_over_configured_api_sender(self, requests_post, get_token):
        get_token.return_value = "token"

        class Response:
            ok = True
            status_code = 200
            headers = {}
            text = ""

            def json(self):
                return {"status": "SUCCESS"}

        requests_post.return_value = Response()

        result = send_sms("+2250749280591", "Bonjour", sender="TEST")

        self.assertTrue(result["success"])
        self.assertEqual(requests_post.call_args.kwargs["json"]["from"], "TEST")

    @override_settings(SMS_API_FROM="0100000000", SMS_API_KEY_BASE64="configured")
    @patch("messaging.sms_gateway._get_token")
    @patch("messaging.sms_gateway.requests.post")
    def test_send_sms_auto_detects_moteur_from_phone_prefix(self, requests_post, get_token):
        get_token.return_value = "token"

        class Response:
            ok = True
            status_code = 200
            headers = {}
            text = ""

            def json(self):
                return {"status": "SUCCESS"}

        requests_post.return_value = Response()

        result = send_sms("+2250749280591", "Bonjour", sender="TEST")

        self.assertTrue(result["success"])
        self.assertEqual(requests_post.call_args.kwargs["json"]["moteur"], "ORANGE")

    @override_settings(SMS_API_FROM="0100000000", SMS_API_KEY_BASE64="configured")
    @patch("messaging.sms_gateway._get_token")
    @patch("messaging.sms_gateway.requests.post")
    def test_send_sms_removes_control_characters_from_message(self, requests_post, get_token):
        get_token.return_value = "token"

        class Response:
            ok = True
            status_code = 200
            headers = {}
            text = ""

            def json(self):
                return {"status": "SUCCESS"}

        requests_post.return_value = Response()

        result = send_sms("+2250749280591", "Bonjour\r\ncher agent\tmerci", sender="TEST")

        self.assertTrue(result["success"])
        self.assertEqual(requests_post.call_args.kwargs["json"]["msg"], "Bonjour cher agent merci")

    @override_settings(SMS_API_FROM="0100000000", SMS_API_KEY_BASE64="configured")
    @patch("messaging.sms_gateway._get_token")
    @patch("messaging.sms_gateway.requests.post")
    def test_send_sms_reports_gateway_json_error_message(self, requests_post, get_token):
        get_token.return_value = "token"

        class Response:
            ok = False
            status_code = 500
            headers = {}
            text = "{\"status\":\"ERROR\",\"message\":\"Échec de l'envoi\"}"

            def json(self):
                return {"status": "ERROR", "message": "Échec de l'envoi"}

        requests_post.return_value = Response()

        result = send_sms("+2250749280591", "Bonjour", sender="TEST")

        self.assertFalse(result["success"])
        self.assertEqual(result["error"], "HTTP 500 : Échec de l'envoi")


class SendSmsNowBillingTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            name="Demo Company",
            balance=Decimal("100.00"),
        )
        self.user = User.objects.create_user(
            username="admin-demo",
            password="password",
            company=self.company,
            role="admin",
        )

    @patch("messaging.services.send_sms_api")
    def test_successful_send_debits_company_and_creates_message(self, send_sms_api):
        send_sms_api.return_value = {"success": True}

        result = send_sms_now(
            user=self.user,
            phone="+2250700000000",
            text="Bonjour",
            title="Test SMS",
        )

        self.assertTrue(result["success"])

        self.company.refresh_from_db()
        self.assertEqual(self.company.balance, Decimal("81.00"))

        self.assertEqual(Message.objects.count(), 1)
        self.assertEqual(CompanyTransaction.objects.count(), 1)

    @patch("messaging.services.send_sms_api")
    def test_api_failure_rolls_back_company_billing(self, send_sms_api):
        send_sms_api.return_value = {"success": False, "error": "API indisponible"}

        result = send_sms_now(
            user=self.user,
            phone="+2250700000000",
            text="Bonjour",
            title="Test SMS",
        )

        self.assertFalse(result["success"])

        self.company.refresh_from_db()
        self.assertEqual(self.company.balance, Decimal("100.00"))

        self.assertEqual(Message.objects.count(), 0)
        self.assertEqual(CompanyTransaction.objects.count(), 0)


class ScheduledSmsProcessingTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            name="Scheduled Company",
            balance=Decimal("100.00"),
        )
        self.user = User.objects.create_user(
            username="scheduled-admin",
            password="password",
            company=self.company,
            role="admin",
        )

    def create_scheduled_message(self):
        return Message.objects.create(
            user=self.user,
            company=self.company,
            title="Scheduled SMS",
            phone="+2250700000000",
            message="Bonjour",
            message_type="simple",
            status="scheduled",
            cost=Decimal("19.00"),
            scheduled_at=timezone.now(),
        )

    @patch("messaging.management.commands.process_scheduled_sms.send_sms_api")
    def test_scheduled_sms_success_debits_company_and_marks_sent(self, send_sms_api):
        msg = self.create_scheduled_message()
        send_sms_api.return_value = {"success": True}

        out = StringIO()
        call_command("process_scheduled_sms", stdout=out)

        msg.refresh_from_db()
        self.company.refresh_from_db()

        self.assertEqual(msg.status, "sent")
        self.assertIsNotNone(msg.sent_at)
        self.assertEqual(self.company.balance, Decimal("81.00"))
        self.assertEqual(CompanyTransaction.objects.count(), 1)
        self.assertIn("1 envoyes", out.getvalue())

    @patch("messaging.management.commands.process_scheduled_sms.send_sms_api")
    def test_scheduled_sms_api_failure_rolls_back_billing_and_marks_failed(self, send_sms_api):
        msg = self.create_scheduled_message()
        send_sms_api.return_value = {"success": False, "error": "API indisponible"}

        out = StringIO()
        call_command("process_scheduled_sms", stdout=out)

        msg.refresh_from_db()
        self.company.refresh_from_db()

        self.assertEqual(msg.status, "failed")
        self.assertEqual(self.company.balance, Decimal("100.00"))
        self.assertEqual(CompanyTransaction.objects.count(), 0)
        self.assertIn("1 echoues", out.getvalue())


class CelerySmsTaskTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            name="Celery Company",
            balance=Decimal("100.00"),
        )
        self.user = User.objects.create_user(
            username="celery-admin",
            password="password",
            company=self.company,
            role="admin",
        )

    def create_pending_message(self, status="pending", sender_name=""):
        return Message.objects.create(
            user=self.user,
            company=self.company,
            title="Celery SMS",
            sender_name=sender_name,
            phone="+2250700000000",
            message="Bonjour",
            message_type="simple",
            status=status,
            cost=Decimal("19.00"),
            scheduled_at=timezone.now() if status == "scheduled" else None,
        )

    @patch("messaging.tasks.send_sms_api")
    def test_send_message_task_success_debits_and_marks_sent(self, send_sms_api):
        msg = self.create_pending_message()
        send_sms_api.return_value = {"success": True}

        result = send_message_task.apply(args=[msg.id], kwargs={"sender": "TEST"}).get()

        msg.refresh_from_db()
        self.company.refresh_from_db()

        self.assertTrue(result["success"])
        self.assertEqual(msg.status, "sent")
        self.assertIsNotNone(msg.sent_at)
        self.assertEqual(self.company.balance, Decimal("81.00"))

    @patch("messaging.tasks.send_sms_api")
    def test_send_message_task_api_failure_rolls_back_billing(self, send_sms_api):
        msg = self.create_pending_message()
        send_sms_api.return_value = {"success": False, "error": "Erreur API gateway"}

        result = send_message_task.apply(args=[msg.id], kwargs={"sender": "TEST"}).get()

        msg.refresh_from_db()
        self.company.refresh_from_db()

        self.assertFalse(result["success"])
        self.assertEqual(msg.status, "failed")
        self.assertEqual(self.company.balance, Decimal("100.00"))

    @patch("messaging.tasks.send_sms_api")
    def test_send_message_task_uses_persisted_sender_name(self, send_sms_api):
        msg = self.create_pending_message(sender_name="ISS")
        send_sms_api.return_value = {"success": True}

        result = send_message_task.apply(args=[msg.id]).get()

        self.assertTrue(result["success"])
        self.assertEqual(send_sms_api.call_args.kwargs["sender"], "ISS")

    @patch("messaging.tasks.send_message_task.delay")
    def test_enqueue_due_scheduled_messages_marks_pending_and_queues_task(self, delay):
        msg = self.create_pending_message(status="scheduled")

        result = enqueue_due_scheduled_messages(limit=10)

        msg.refresh_from_db()

        self.assertEqual(result["queued"], 1)
        self.assertEqual(msg.status, "pending")
        delay.assert_called_once_with(msg.id)


class MessagingViewValidationTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            name="View Company",
            balance=Decimal("100.00"),
        )
        self.user = User.objects.create_user(
            username="view-admin",
            password="password",
            company=self.company,
            role="admin",
        )
        self.sender = Sender.objects.create(
            company=self.company,
            created_by=self.user,
            name="TEST",
            status="approved",
        )
        self.client.force_login(self.user)

    def test_send_sms_rejects_invalid_schedule_date(self):
        response = self.client.post(reverse("send_sms"), {
            "title": "Test SMS",
            "sender": str(self.sender.id),
            "phone": "+2250700000000",
            "message": "Bonjour",
            "scheduled_at": "date-invalide",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Date de programmation invalide.")
        self.assertContains(response, "Test SMS")
        self.assertContains(response, "+2250700000000")
        self.assertEqual(Message.objects.count(), 0)

    @override_settings(SMS_API_KEY_BASE64="configured")
    def test_send_sms_queues_pending_message(self):
        response = self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "+2250700000000",
                "message": "Bonjour",
            },
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))
        msg = Message.objects.get()

        self.assertEqual(len(django_messages), 1)
        self.assertEqual(django_messages[0].tags, "success")
        self.assertIn("mis en file", str(django_messages[0]))
        self.assertEqual(msg.status, "pending")
        self.assertEqual(msg.phone, "0700000000")

    @override_settings(SMS_API_KEY_BASE64="configured")
    @patch("messaging.views.queue_sms_send")
    def test_send_sms_send_now_ignores_hidden_schedule_date(self, queue_sms_send):
        queue_sms_send.return_value = {"success": True}
        future_date = (timezone.now() + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M")

        self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "+2250700000000",
                "message": "Bonjour",
                "send_action": "send_now",
                "scheduled_at": future_date,
            },
        )

        queue_sms_send.assert_called_once()
        self.assertEqual(Message.objects.count(), 0)

    def test_send_sms_schedule_action_requires_date(self):
        response = self.client.post(reverse("send_sms"), {
            "title": "Test SMS",
            "sender": str(self.sender.id),
            "phone": "+2250700000000",
            "message": "Bonjour",
            "send_action": "schedule",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Veuillez choisir une date de programmation.")
        self.assertEqual(Message.objects.count(), 0)

    def test_send_sms_scheduled_message_stores_selected_sender(self):
        future_date = (timezone.now() + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M")

        self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "+2250700000000",
                "message": "Bonjour",
                "send_action": "schedule",
                "scheduled_at": future_date,
            },
        )

        msg = Message.objects.get()
        self.assertEqual(msg.status, "scheduled")
        self.assertEqual(msg.sender_name, "TEST")

    @override_settings(SMS_API_KEY_BASE64="configured")
    @patch("messaging.views.queue_sms_send")
    def test_send_sms_passes_selected_sender_to_queue_service(self, queue_sms_send):
        queue_sms_send.return_value = {"success": True}
        self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "+2250700000000",
                "message": "Bonjour",
            },
        )

        self.assertEqual(queue_sms_send.call_args.kwargs["sender"], "TEST")

    @override_settings(SMS_API_KEY_BASE64="", SMS_API_CLIENT_ID="", SMS_API_CLIENT_SECRET="")
    @patch("messaging.views.queue_sms_send")
    def test_send_sms_missing_api_key_stops_before_processing_numbers(self, queue_sms_send):
        response = self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "+2250700000000;+2250700000001",
                "message": "Bonjour",
            },
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))

        queue_sms_send.assert_not_called()
        self.assertEqual(len(django_messages), 1)
        self.assertEqual(django_messages[0].tags, "error")
        self.assertIn("Configuration SMS incomplète", str(django_messages[0]))

    @override_settings(SMS_API_KEY_BASE64="configured")
    def test_send_sms_only_invalid_numbers_shows_single_error(self):
        response = self.client.post(
            reverse("send_sms"),
            {
                "title": "Test SMS",
                "sender": str(self.sender.id),
                "phone": "12345",
                "message": "Bonjour",
            },
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))

        self.assertEqual(Message.objects.count(), 0)
        self.assertEqual(len(django_messages), 1)
        self.assertEqual(django_messages[0].tags, "error")
        self.assertIn("Aucun SMS envoyé. Numéros invalides", str(django_messages[0]))

    def test_campaign_upload_rejects_invalid_group_without_creating_campaign(self):
        response = self.client.post(reverse("campaign_upload"), {
            "title": "Campagne test",
            "sender": str(self.sender.id),
            "contact_group": "9999",
            "message": "Bonjour",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Groupe de contacts invalide.")
        self.assertContains(response, "Campagne test")
        self.assertEqual(Campaign.objects.count(), 0)

    @override_settings(SMS_API_KEY_BASE64="configured")
    @patch("messaging.views.enqueue_pending_messages.delay")
    def test_campaign_upload_queues_one_background_task_for_excel_contacts(self, delay):
        upload = make_campaign_excel([
            ["0777186049", "CH"],
            ["0768944994", "Harouna"],
            ["12345", "Invalid"],
        ])

        response = self.client.post(
            reverse("campaign_upload"),
            {
                "title": "Campagne test",
                "sender": str(self.sender.id),
                "message": "Bonjour {{name}}",
                "file": upload,
            },
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))
        pending_ids = list(Message.objects.filter(status="pending").order_by("id").values_list("id", flat=True))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(pending_ids), 2)
        self.assertEqual(Message.objects.filter(phone="0777186049").count(), 1)
        delay.assert_called_once_with(pending_ids)
        self.assertTrue(any("2 SMS mis en file" in str(message) for message in django_messages))

    def test_contact_group_stats_returns_contact_preview_for_campaign(self):
        group = ContactGroup.objects.create(
            name="DT",
            company=self.company,
            owner=self.user,
        )
        Contact.objects.create(
            company=self.company,
            user=self.user,
            group=group,
            phone="+2250777186049",
            name="CH",
        )
        Contact.objects.create(
            company=self.company,
            user=self.user,
            group=group,
            phone="+2250768944994",
            name="Harouna COULIBALY",
        )

        response = self.client.get(reverse("contact_group_stats"), {"group_id": group.id})
        payload = response.json()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["total"], 2)
        self.assertEqual(payload["valid"], 2)
        self.assertEqual(len(payload["contacts"]), 2)
        self.assertEqual(payload["contacts"][0]["phone"], "0777186049")

    def test_download_campaign_model_excel(self):
        response = self.client.get(reverse("download_model"))
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(response.status_code, 200)
        self.assertIn("modele_import_campagne_sms.xlsx", response["Content-Disposition"])
        self.assertEqual([sheet.cell(1, col).value for col in range(1, 3)], ["phone", "name"])
        self.assertEqual(sheet.cell(2, 1).value, "0777186049")

    def test_message_history_csv_export_uses_current_filters(self):
        Message.objects.create(
            user=self.user,
            company=self.company,
            title="Export CSV",
            phone="+2250700000000",
            message="Bonjour export",
            message_type="simple",
            status="sent",
            cost=Decimal("19.00"),
            sent_at=timezone.now(),
        )
        Message.objects.create(
            user=self.user,
            company=self.company,
            title="Scheduled",
            phone="+2250500000000",
            message="Bonjour planifie",
            message_type="simple",
            status="scheduled",
            cost=Decimal("19.00"),
            scheduled_at=timezone.now(),
        )

        response = self.client.get(reverse("message_history"), {
            "status": "sent",
            "export": "csv",
        })

        content = response.content.decode("utf-8")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("historique_messages.csv", response["Content-Disposition"])
        self.assertIn("Export CSV", content)
        self.assertNotIn("Scheduled", content)

    def test_message_search_api_returns_title(self):
        Message.objects.create(
            user=self.user,
            company=self.company,
            title="Recherche titre",
            phone="+2250700000000",
            message="Bonjour recherche",
            message_type="simple",
            status="sent",
            cost=Decimal("19.00"),
            sent_at=timezone.now(),
        )

        response = self.client.get(reverse("message_search"), {"q": "titre"})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["results"][0]["title"], "Recherche titre")

    def test_simple_user_cannot_manage_senders(self):
        simple_user = User.objects.create_user(
            username="simple-user",
            password="password",
            company=self.company,
            role="user",
        )
        self.client.force_login(simple_user)

        response = self.client.get(reverse("senders_list"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "messaging/forbidden.html")

    def test_sender_create_duplicate_shows_user_error(self):
        response = self.client.post(
            reverse("sender_create"),
            {"name": "test"},
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Sender.objects.filter(company=self.company, name="TEST").count(), 1)
        self.assertTrue(any("existe déjà" in str(message) for message in django_messages))

    def test_sender_edit_duplicate_shows_user_error(self):
        other_sender = Sender.objects.create(
            company=self.company,
            created_by=self.user,
            name="OTHER",
            status="pending",
        )

        response = self.client.post(
            reverse("sender_edit", args=[other_sender.id]),
            {"name": "test"},
            follow=True,
        )

        django_messages = list(get_messages(response.wsgi_request))
        other_sender.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(other_sender.name, "OTHER")
        self.assertTrue(any("existe déjà" in str(message) for message in django_messages))
