from io import BytesIO

import openpyxl
import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import ContactGroup, Contact
from .phone_numbers import is_valid_phone, normalize_phone, phone_lookup_values

from django.core.paginator import Paginator
from django.db.models import Q

from contacts.utils import can_access_group, is_company_admin


logger = logging.getLogger(__name__)
PHONE_FORMAT_HELP = "Format attendu : 0XXXXXXXXX."

######################
# UTILS SÉCURITÉ
######################


def can_access_contact(user, contact):
    if is_company_admin(user):
        return contact.company == user.company
    return contact.company == user.company and contact.user == user


######################
# GROUPES
######################

@login_required
def contact_groups_list(request):

    user = request.user

    if is_company_admin(user):
        groups = ContactGroup.objects.filter(
            company=user.company
        ).select_related("owner")

    else:
        groups = ContactGroup.objects.filter(
            company=user.company,
            owner=user
        ).select_related("owner")

    return render(request, "contacts/groups_list.html", {
        "groups": groups
    })

# group create
@login_required
def contact_group_create(request):

    if not request.user.company_id:
        return render(request, "messaging/forbidden.html")

    if request.method == "POST":

        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            return render(request, "contacts/group_create.html", {
                "error": "Le nom du groupe est obligatoire."
            })

        if ContactGroup.objects.filter(company=request.user.company, name=name).exists():
            return render(request, "contacts/group_create.html", {
                "error": "Un groupe avec ce nom existe déjà.",
                "name": name,
                "description": description,
            })

        try:
            ContactGroup.objects.create(
                name=name,
                description=description,
                owner=request.user,
                company=request.user.company
            )
        except ValidationError:
            logger.exception("Erreur validation création groupe contacts")
            return render(request, "contacts/group_create.html", {
                "error": "Impossible de créer ce groupe avec ces informations.",
                "name": name,
                "description": description,
            })

        messages.success(request, "Groupe créé avec succès.")
        return redirect("contact_groups_list")

    return render(request, "contacts/group_create.html")


# group detail




@login_required
def contact_group_detail(request, group_id):

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    # ==============================
    # 🔍 RECHERCHE
    # ==============================
    search = request.GET.get("q", "").strip()

    # ==============================
    # 📊 BASE QUERY (SECURITE)
    # ==============================
    if is_company_admin(request.user):
        contacts = group.contacts.filter(company=request.user.company)
    else:
        contacts = group.contacts.filter(
            company=request.user.company,
            user=request.user
        )

    # ==============================
    # 🔎 FILTRE RECHERCHE
    # ==============================
    if search:
        contacts = contacts.filter(
            Q(phone__icontains=search) |
            Q(name__icontains=search) |
            Q(email__icontains=search)
        )

    # ==============================
    # 📄 PAGINATION
    # ==============================
    contacts = contacts.order_by("-created_at")

    paginator = Paginator(contacts, 10)  # 10 contacts / page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ==============================
    # 🎯 RENDER
    # ==============================
    return render(request, "contacts/group_detail.html", {
        "group": group,
        "contacts": page_obj,   # 🔥 IMPORTANT
        "search": search,       # 🔥 pour garder la recherche dans input
    })



# group edit
@login_required
def contact_group_edit(request, group_id):

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            return render(request, "contacts/group_edit.html", {
                "group": group,
                "error": "Le nom du groupe est obligatoire."
            })

        duplicate = ContactGroup.objects.filter(
            company=request.user.company,
            name=name
        ).exclude(id=group.id).exists()

        if duplicate:
            group.name = name
            group.description = description
            return render(request, "contacts/group_edit.html", {
                "group": group,
                "error": "Un groupe avec ce nom existe déjà."
            })

        try:
            group.name = name
            group.description = description
            group.save()
        except ValidationError:
            logger.exception("Erreur validation modification groupe contacts")
            return render(request, "contacts/group_edit.html", {
                "group": group,
                "error": "Impossible d'enregistrer ce groupe avec ces informations."
            })

        messages.success(request, "Groupe modifié avec succès.")
        return redirect("contact_groups_list")

    return render(request, "contacts/group_edit.html", {
        "group": group
    })


@login_required
def contact_group_delete(request, group_id):

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    if request.method != "POST":
        messages.error(request, "Confirmation requise pour supprimer un groupe.")
        return redirect("contact_groups_list")

    group.delete()
    messages.success(request, "Groupe supprimé avec succès.")

    return redirect("contact_groups_list")

# group stats
@login_required
def contact_group_stats(request):

    group_id = request.GET.get("group_id")

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return JsonResponse({"error": "Forbidden"}, status=403)

    contacts = Contact.objects.filter(
        group=group,
        company=request.user.company
    )

    if not is_company_admin(request.user):
        contacts = contacts.filter(user=request.user)

    contacts = contacts.order_by("name", "phone")
    total = contacts.count()
    valid = sum(1 for contact in contacts if is_valid_phone(contact.phone))
    invalid = total - valid
    preview_contacts = [
        {
            "phone": normalize_phone(contact.phone),
            "name": contact.name or "",
            "valid": is_valid_phone(contact.phone),
        }
        for contact in contacts[:10]
    ]

    return JsonResponse({
        "total": total,
        "valid": valid,
        "invalid": invalid,
        "contacts": preview_contacts,
    })


######################
# CONTACTS
######################

@login_required
def contact_create(request, group_id):

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    if request.method == "POST":

        phone = normalize_phone(request.POST.get("phone"))
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip() or None

        if not is_valid_phone(phone):
            messages.error(request, f"Numéro invalide. {PHONE_FORMAT_HELP}")
            return redirect("contact_group_detail", group_id=group.id)

        if Contact.objects.filter(
            phone__in=phone_lookup_values(phone),
            user=request.user,
        ).exists():
            messages.error(request, "Ce numéro existe déjà dans vos contacts.")
            return redirect("contact_group_detail", group_id=group.id)

        try:
            Contact.objects.create(
                group=group,
                phone=phone,
                user=request.user,
                name=name,
                email=email,
                company=request.user.company,
            )
            messages.success(request, "Contact ajouté avec succès.")
        except ValidationError:
            logger.exception("Erreur validation création contact")
            messages.error(request, "Impossible d'ajouter ce contact avec ces informations.")

    return redirect("contact_group_detail", group_id=group.id)


@login_required
def contacts_import_excel(request, group_id):

    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    results = None
    rows_data = []

    if request.method == "POST":

        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Aucun fichier sélectionné")
            return redirect("contact_group_detail", group_id=group.id)

        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
        except Exception:
            messages.error(request, "Fichier Excel illisible.")
            return redirect("contacts_import_excel", group_id=group.id)

        sheet = workbook.active

        total_rows = 0
        created = 0
        duplicates = []
        invalid = []
        errors = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            phone = ""
            name = ""
            email = ""
            status = "error"

            try:

                phone = normalize_phone(row[0])

                if not phone:
                    continue

                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                total_rows += 1

                # ❌ NUMERO INVALIDE
                if not is_valid_phone(phone):
                    invalid.append(phone)
                    status = "invalid"

                # 🔁 DUPLICATE
                elif Contact.objects.filter(
                    phone__in=phone_lookup_values(phone),
                    user=request.user
                ).exists():

                    duplicates.append(phone)
                    status = "duplicate"

                else:

                    try:
                        Contact.objects.create(
                            group=group,
                            phone=phone,
                            name=name or "",
                            email=email or None,
                            company=request.user.company,
                            user=request.user
                        )

                        created += 1
                        status = "imported"

                    except Exception as e:
                        status = "error"
                        errors.append(phone)
                        logger.exception("Erreur import contact %s", phone)

            except Exception as e:
                # 💥 CAPTURE ERREUR DJANGO (full_clean etc.)
                errors.append(phone or "ligne inconnue")
                status = "error"
                logger.exception("Erreur lecture ligne import contacts")

            rows_data.append({
                "phone": phone,
                "name": name,
                "email": email,
                "status": status
            })

        results = {
            "total_rows": total_rows,
            "created": created,
            "duplicates": duplicates,
            "invalid": invalid,
            "errors": errors
        }

        messages.success(
            request,
            f"{created} importés | {len(duplicates)} doublons | {len(invalid)} invalides"
        )

    return render(request, "contacts/import_excel.html", {
        "group": group,
        "results": results,
        "rows_data": rows_data
    })


@login_required
def contacts_download_model_excel(request, group_id):
    group = get_object_or_404(
        ContactGroup,
        id=group_id,
        company=request.user.company
    )

    if not can_access_group(request.user, group):
        return render(request, "messaging/forbidden.html")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"

    headers = ["phone", "name", "email"]
    examples = [
        ["0777186049", "CH", "contact@example.com"],
        ["0768944994", "Harouna COULIBALY", "harouna@example.com"],
    ]

    sheet.append(headers)
    for row in examples:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="009846")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 34
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="modele_import_contacts.xlsx"'

    return response



@login_required
def contact_edit(request, contact_id):

    contact = get_object_or_404(
        Contact,
        id=contact_id,
        company=request.user.company
    )

    if not can_access_contact(request.user, contact):
        return render(request, "messaging/forbidden.html")

    if request.method == "POST":
        phone = normalize_phone(request.POST.get("phone"))

        if not is_valid_phone(phone):
            messages.error(request, f"Numéro invalide. {PHONE_FORMAT_HELP}")
            return redirect("contact_edit", contact_id=contact.id)

        if Contact.objects.filter(
            user=request.user,
            phone__in=phone_lookup_values(phone),
        ).exclude(id=contact.id).exists():
            messages.error(request, "Ce numéro existe déjà dans vos contacts.")
            return redirect("contact_edit", contact_id=contact.id)

        contact.phone = phone
        contact.name = (request.POST.get("name") or "").strip()
        contact.email = (request.POST.get("email") or "").strip() or None
        try:
            contact.save()
            messages.success(request, "Contact modifié avec succès.")
        except ValidationError:
            logger.exception("Erreur validation modification contact")
            messages.error(request, "Impossible d'enregistrer ce contact avec ces informations.")
            return redirect("contact_edit", contact_id=contact.id)

        return redirect("contact_group_detail", group_id=contact.group.id)

    return render(request, "contacts/contact_edit.html", {
        "contact": contact
    })


@login_required
def contact_delete(request, contact_id):

    contact = get_object_or_404(
        Contact,
        id=contact_id,
        company=request.user.company
    )

    if not can_access_contact(request.user, contact):
        return render(request, "messaging/forbidden.html")

    if request.method != "POST":
        messages.error(request, "Confirmation requise pour supprimer un contact.")
        return redirect("contact_group_detail", group_id=contact.group.id)

    group_id = contact.group.id
    contact.delete()
    messages.success(request, "Contact supprimé avec succès.")
    return redirect("contact_group_detail", group_id=group_id)
